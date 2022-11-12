import os
import re
import sys
import csv
import json
import glob
import time
import codecs
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
import string

# Suppress as many warnings as possible
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
from tensorflow.python.util import deprecation
deprecation._PRINT_DEPRECATION_WARNINGS = False
import tensorflow as tf
tf.compat.v1.logging.set_verbosity(tf.compat.v1.logging.ERROR)

from ckiptagger import data_utils, construct_dictionary, WS, POS, NER

#from pywordseg import *

def replacing(candidates, text):
    for candidate in candidates:
        text = text.replace(*candidate)
    return text

def get_titles(limit=1000):
    i = 1
    di = {}
    for i1 in string.ascii_uppercase:
        di[i1] = i
        if i == limit:
            return di
        i += 1
    for i1 in string.ascii_uppercase:
        for i2 in string.ascii_uppercase:
            di[i1 + i2] = i
            if i == limit:
                return di
            i += 1
    for i1 in string.ascii_uppercase:
        for i2 in string.ascii_uppercase:
            for i3 in string.ascii_uppercase:
                di[i1 + i2 + i3] = i
                if i == limit:
                    return di
                i += 1
    


class GUIDemo(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.grid()
        self.createWidgets()
        self.fin = None
        self.gex = re.compile(r"\((\w+)\)")
        # self.replacements = list(csv.reader(open("convert.tsv"), delimiter='\t'))
        # self.pos_list = list(csv.reader(open("statistics.tsv"), delimiter='\t'))
        self.replacements = [['。', '.'], ['，', ','], ['、', ','], ['；', ';'], ['：', ':'], ['「', '"'], ['」', '"'], ['『', '"'], ['』', '"'], ['（', '('], ['）', ')'], ['【', '('], ['】', ')'], ['？', '?'], ['！', '!'], ['──', '-'], ['─', '-'], ['……', '…'], ['—', '-'], ['＃', '#'], ['＄', '$'], ['％', '%'], ['＆', '&'], ['＊', '*'], ['＋', '+'], ['／', '/'], ['＜', '<'], ['＝', '='], ['＞', '>'], ['^', '^'], ['﹍', '_'], ['～', '~'], ['\u3000', ' ']]
        #self.pos_list = [['A', 'c_A'], ['Caa', 'c_Caa'], ['Cab', 'c_Cab'], ['Cba', 'c_Cba'], ['Cbb', 'c_Cbb'], ['Da', 'c_Da'], ['Dfa', 'c_Dfa'], ['Dfb', 'c_Dfb'], ['Di', 'c_Di'], ['Dk', 'c_Dk'], ['D', 'c_D'], ['Na', 'c_Na'], ['Nb', 'c_Nb'], ['Nc', 'c_Nc'], ['Ncd', 'c_Ncd'], ['Nd', 'c_Nd'], ['Neu', 'c_Neu'], ['Nes', 'c_Nes'], ['Nep', 'c_Nep'], ['Neqa', 'c_Neqa'], ['Neqb', 'c_Neqb'], ['Nf', 'c_Nf'], ['Ng', 'c_Ng'], ['Nh', 'c_Nh'], ['Nv', 'c_Nv'], ['I', 'c_I'], ['P', 'c_P'], ['T', 'c_T'], ['VA', 'c_VA'], ['VAC', 'c_VAC'], ['VB', 'c_VB'], ['VC', 'c_VC'], ['VCL', 'c_VCL'], ['VD', 'c_VD'], ['VE', 'c_VE'], ['VF', 'c_VF'], ['VG', 'c_VG'], ['VH', 'c_VH'], ['VHC', 'c_VHC'], ['VI', 'c_VI'], ['VJ', 'c_VJ'], ['VK', 'c_VK'], ['VL', 'c_VL'], ['V_2', 'c_V_2'], ['DE', 'c_DE'], ['SHI', 'c_SHI'], ['FW', 'c_FW'], ['COMMACATEGORY', 'c_comma'], ['DASHCATEGORY', 'c_dash'], ['ETCCATEGORY', 'c_etc'], ['EXCLAMATIONCATEGORY', 'c_exclam'], ['PARENTHESISCATEGORY', 'c_parenth'], ['PAUSECATEGORY', 'c_pause'], ['PERIODCATEGORY', 'c_period'], ['QUESTIONCATEGORY', 'c_qmark'], ['COLONCATEGORY', 'c_colon'], ['SEMICOLONCATEGORY', 'c_semic'], ['SPCHANGECATEGORY', 'c_sp']]
        #self.comma_dict = {'COMMACATEGORY': 'c_comma', 'DASHCATEGORY': 'c_dash', 'ETCCATEGORY': 'c_etc', 'EXCLAMATIONCATEGORY': 'c_exclam', 'PARENTHESISCATEGORY': 'c_parenth', 'PAUSECATEGORY': 'c_pause', 'PERIODCATEGORY': 'c_period', 'QUESTIONCATEGORY': 'c_qmark', 'COLONCATEGORY': 'c_colon', 'SEMICOLONCATEGORY': 'c_semic', 'SPCHANGECATEGORY': 'c_sp'}
        self.column_titles = get_titles(1000)
        self.wordseg_root = None
        self.wordseg_dict = None
        self.wordseg = None
#        self.CkipText["text"] = "--- Pywordseg 初始化中 ---"
        self.CkipText["text"] = "--- CkipTagger 尚未初始化 ---"
        # try:
        #     self.ws = WS("./data")
        #     #self.wordseg = Wordseg(batch_size=64, device="cpu", embedding='w2v', elmo_use_cuda=False, mode="TW")
        #     self.CkipText["text"] = "--- CkipTagger 初始化成功 ---"
        #     #self.CkipText["text"] = "--- Pywordseg 初始化成功 ---"
        # except:
        #     self.CkipText["text"] = "Error: CkipTagger 初始化失敗"
        #     #self.CkipText["text"] = "Error: Pywordseg 初始化失敗"

    def set_wordseg_dict(self):
        if self.wordseg_dict is not None:
            self.wordseg_dict = None
            self.CkipText2["text"] = "--- 清除Pywordseg字典檔 ---"
        path = filedialog.askopenfilename(initialdir=os.getcwd())
        if len(path) == 0:
            return
        self.wordseg_root = path
        try:
            assert os.path.exists(path)
            path = path.replace("/", "\\")
            self.wordseg_dict = construct_dictionary(json.load(open(path)))
            self.CkipText["text"] = "--- 設定Pywordseg字典檔為 ---"
            self.CkipText2["text"] = path
        except:
            self.CkipText["text"] = "Warning: 字典檔讀取錯誤"

    def download_model(self):
        data_utils.download_data("./")

    def detect_coding(self, filename):
        candidates = ['utf-8', 'cp950', 'big5-hkscs', 'utf-16']
        for candidate in candidates:
            try:
                f = codecs.open(filename, encoding=candidate)
                _ = f.readline()
                return candidate
            except:
                continue
        self.displayText2["text"] = f"File `{os.path.split(filename)[1]}` encoding not by UTF-8 or ANSI."
        return None

    def choose(self):
        self.fin = filedialog.askopenfilename()
        if self.fin == "":
            return
        basename, filename = os.path.split(self.fin)
        outname = os.path.join("result", "fmt_"+filename)
        self.displayText["text"] = "Input file: "+filename
        #self.displayText2["text"] = "Output to: "+outname

    def choose_files(self):
        self.fin = filedialog.askopenfilenames(filetypes=[("Text Files", "*.xlsx")])
        if len(self.fin) == 0:
            return
        filenames = [os.path.split(fin)[1] for fin in self.fin]
        #outnames = [os.path.join("result", "fmt_"+filename) for filename in filenames]
        self.displayText["text"] = "---- Input File ----\n"+"\n".join(filenames)
        self.displayText2["text"] = ""

    def choose_dir(self):
        mydir = filedialog.askdirectory(initialdir=os.getcwd())
        if len(mydir) == 0:
            return
        self.fin = glob.glob(os.path.join(mydir, "*.xlsx"))
        filenames = [os.path.split(fin)[1] for fin in self.fin]
        #outnames = [os.path.join("result", "fmt_"+filename) for filename in filenames]
        self.displayText["text"] = "---- Input File ----\n" + "\n".join(filenames)
        self.displayText2["text"] = ""

    
    def choose_files_txt(self):
        self.fin = filedialog.askopenfilenames(filetypes=[("Text Files", "*.txt")])
        if len(self.fin) == 0:
            return
        filenames = [os.path.split(fin)[1] for fin in self.fin]
        #outnames = [os.path.join("result", "fmt_"+filename) for filename in filenames]
        self.displayText["text"] = "---- Input File ----\n"+"\n".join(filenames)
        self.displayText2["text"] = ""

    def choose_dir_txt(self):
        mydir = filedialog.askdirectory(initialdir=os.getcwd())
        if len(mydir) == 0:
            return
        self.fin = glob.glob(os.path.join(mydir, "*.txt"))
        filenames = [os.path.split(fin)[1] for fin in self.fin]
        #outnames = [os.path.join("result", "fmt_"+filename) for filename in filenames]
        self.displayText["text"] = "---- Input File ----\n" + "\n".join(filenames)
        self.displayText2["text"] = ""
    
    def choose_model_txt(self):
        mydir = filedialog.askdirectory(initialdir=os.getcwd())
        # self.CkipText["text"] = "--- CkipTagger 初始化中 ---"
        if len(mydir) == 0:
            return
        try:
            self.ws = WS(mydir)
            #self.wordseg = Wordseg(batch_size=64, device="cpu", embedding='w2v', elmo_use_cuda=False, mode="TW")
            self.CkipText["text"] = "--- CkipTagger 初始化成功 ---"
            self.wordseg_button['state'] = 'disabled'
        except:
            self.CkipText["text"] = "Error: CkipTagger 初始化失敗"

    def assert_msg(self, cond, text):
        if cond:
            return True
        else:
            self.displayText2["text"] = text
            return False

    def process(self):
        start = time.time()
        if type(self.fin) == list:
            files = self.fin
        elif type(self.fin) == tuple:
            files = self.fin
        elif type(self.fin) == str:
            files = [self.fin]
        else:
            self.displayText2["text"] = "Unknown filetype " + str(type(self.fin)) + ": " + str(self.fin)
            return
        done_seg = False #(self.done_variable.get() == "Yes")
        if self.ws is None and not done_seg:
 #       if self.wordseg is None and not done_seg:
            #self.displayText2["text"] = "Error: Pywordseg 初始化未完成或失敗"
            self.displayText2["text"] = "Error: CkipTagger WS 初始化未完成或失敗"
            return
        donefile = []
        column = self.column_entry.get().upper()
        # done_seg = False
        if not self.assert_msg(column in self.column_titles, "Unknown column " + column):
            return
        column_idx = self.column_titles[column]
        for idx, filename in enumerate(files):
            if filename[-5:] == '.xlsx':
                wb = load_workbook(filename)
                for sheetnames in wb.sheetnames:
                    ws = wb[sheetnames]
                    if not self.assert_msg(column_idx <= ws.max_column, "Invalid column index (%d > max column %d)" % (column_idx, ws.max_column)):
                        return
                    ws.insert_cols(column_idx+1)
                    tmp_txts = []
                    tmp_rows = []
                    for row in range(1, ws.max_row + 1):
                        text = ws.cell(row, column_idx).value
                        if text is None:
                            continue
                        text = str(text)
                        #if not done_seg:
                        #text = self.wordseg.cut([text], merge_dict=self.wordseg_dict)[0]
                        text = text.replace('\r\n', '')
                        # text = re.sub(r"(\r\n)+", '\r\n', text)
                        # text = re.sub(r"\n+", '\n', text)
                        text = text.replace('\n', '')
                        tmp_txts.append(text)
                        tmp_rows.append(row)
                    seg_txts = self.ws(tmp_txts)
                    # seg_txts = self.wordseg.cut(tmp_txts, merge_dict=self.wordseg_dict)
                    for row, text in zip(tmp_rows, seg_txts):
                        text = ' '.join(text)
                        text = replacing(self.replacements, text)
                        #text = re.sub(r"\(\w+\)", '', text)
                        ws.cell(row, column_idx+1).value = text
                basename, filename = os.path.split(filename)
                os.makedirs(os.path.join(basename, 'result'), exist_ok=True)
                outname = os.path.join(basename, 'result', 'fmt_'+filename)
                wb.save(outname)
                donefile.append(os.path.join('result', 'fmt_'+filename))
                self.displayText["text"] = "---- Output File ----\n" + '\n'.join(donefile)
            elif filename[-4:] == '.txt':
                codec = self.detect_coding(filename)
                if codec is None:
                    return
                fin = open(filename, 'r', encoding=codec)
                basename, filename = os.path.split(filename)
                os.makedirs(os.path.join(basename, 'result'), exist_ok=True)
                outname = os.path.join(basename, 'result', 'fmt_'+filename)
                fout = open(outname, 'w', encoding='utf-8-sig', newline='\n')

                tmp_txts = []
                #tmp_rows = []
                for line in fin:
                    #if not done_seg:
                    #line = self.wordseg.cut([line], merge_dict=self.wordseg_dict)[0]
                    line = line.replace('\r\n', '')
                    # line = re.sub(r"(\r\n)+", '\r\n', line)
                    # line = re.sub(r"\n+", '\n', line)
                    # line = line.replace('\n\n', '\n')
                    line = line.replace('\n', '')
                    tmp_txts.append(line)
                    #tmp_rows.append(row)
#                seg_txts = self.wordseg.cut(tmp_txts, merge_dict=self.wordseg_dict)
                seg_txts = self.ws(tmp_txts)
                for line in seg_txts:
                    line = ' '.join(line)
                    line = replacing(self.replacements, line)
                    line = re.sub(r"(: | : )",' : ',line)
                    fout.write(line.strip()+'\n')
                donefile.append(os.path.join('result', 'fmt_'+filename))
                self.displayText["text"] = "---- Output File ----\n" + '\n'.join(donefile)
        period = time.time() - start
        self.displayText2["text"] = "處理完成 (in %.3f s)"%period
        
    def createWidgets(self):
        row = 1
        self.topText = Label(self, width=20, height=1)
        self.topText["text"] = "\u3000"
        self.topText.grid(row=row, column=0, columnspan=5)
        row += 1
        
        self.gap = Label(self, width=5, height=1)
        self.gap.grid(row=row, column=0)

        self.wordseg_button = Button(self, width=20, height=1)
        # self.wordseg_button["text"] = "下載 CkipTagger 模型檔" # v1 小路
        self.wordseg_button["text"] = "導入 CkipTagger 模型檔" 
        #self.wordseg_button["text"] = "導入字典檔"
        self.wordseg_button.grid(row=row, column=1, columnspan=2)
        self.wordseg_button["command"] = self.choose_model_txt
        #self.wordseg_button["command"] = self.set_wordseg_dict
        row += 1

        self.CkipText = Label(self)
        self.CkipText["text"] = "--- CkipTagger 尚未初始化 ---"
        #self.CkipText["text"] = "--- Pywordseg 尚未初始化 ---"
        self.CkipText.grid(row=row, column=0, columnspan=7)
        row += 1
        self.CkipText2 = Label(self)
        self.CkipText2["text"] = ""
        self.CkipText2.grid(row=row, column=0, columnspan=7)
        row += 1

        self.HeadText2 = Label(self)
        self.HeadText2["text"] = "__________ 純文字文件 __________"
        self.HeadText2.grid(row=row, column=0, columnspan=7)
        row += 1

        self.register_button3 = Button(self, width=9, height=1)
        self.register_button3["text"] = "開啟文字檔"
        self.register_button3.grid(row=row, column=1)
        self.register_button3["command"] = self.choose_files_txt


        self.register_button4 = Button(self, width=9, height=1)
        self.register_button4["text"] = "開啟資料夾"
        self.register_button4.grid(row=row, column=2)
        self.register_button4["command"] = self.choose_dir_txt

        self.gap3 = Label(self, width=5, height=1)
        self.gap3.grid(row=row, column=3)
        row += 1


        self.HeadText = Label(self)
        self.HeadText["text"] = "__________ EXCEL文件 __________"
        self.HeadText.grid(row=row, column=0, columnspan=7)
        row += 1

        self.register_button1 = Button(self, width=9, height=1)
        self.register_button1["text"] = "開啟xlsx檔"
        self.register_button1.grid(row=row, column=1)
        self.register_button1["command"] = self.choose_files


        self.register_button2 = Button(self, width=9, height=1)
        self.register_button2["text"] = "開啟資料夾"
        self.register_button2.grid(row=row, column=2)
        self.register_button2["command"] = self.choose_dir

        self.gap2 = Label(self, width=5, height=1)
        self.gap2.grid(row=row, column=3)
        row += 1

        self.column_label = Label(self, text='輸入欲處理欄位')
        self.column_label.grid(row=row, column=1)
        self.column_variable = StringVar(self, value='B')
        self.column_entry = Entry(self, width=10, textvariable=self.column_variable)
        self.column_entry.grid(row=row, column=2)
        row += 1
        

        #self.done_label = Label(self, text='是否已斷詞過')
        #self.done_label.grid(row=row, column=1)
        #self.done_variable = StringVar(self, value='No')
        #self.done_segmt = OptionMenu(self, self.done_variable, "Yes", "No")
        #self.done_segmt.grid(row=row, column=2)
        #row += 1
                
        self.displayText3 = Label(self)
        self.displayText3["text"] = ""
        self.displayText3.grid(row=row, column=0, columnspan=7)
        row += 1
        
        self.unlock_button = Button(self, width=20, height=1)
        self.unlock_button["text"] = "開始處理"
        self.unlock_button.grid(row=row, column=1, columnspan=2)
        self.unlock_button["command"] = self.process
        row += 1
        
        self.displayText = Label(self)
        self.displayText["text"] = ""
        self.displayText.grid(row=row, column=0, columnspan=7)
        row += 1

        self.displayText2 = Label(self)
        self.displayText2["text"] = ""
        self.displayText2.grid(row=row, column=0, columnspan=7)
        row += 1
        
if __name__ == '__main__':
    root = Tk(className=" CkipTagger GUI Tool ")
    #root = Tk(className=" Pywordseg GUI Tool ")
    app = GUIDemo(master=root)
    app.mainloop()
